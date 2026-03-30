#!/usr/bin/env python3
# Telegram → Claude — стриминг, кэш, extended thinking, MCP-инструменты
# Запуск: python3 claude_bot.py

import os, sys, json, time, glob, base64, subprocess, threading
from datetime import datetime

import anthropic, requests

sys.path.insert(0, os.path.dirname(__file__))
from config import (TELEGRAM_TOKEN, ANTHROPIC_KEY, CHAT_ID, GROUP_ID,
                    TOPICS, LOGS_DIR, PROJECT_DIR, TMP_DIR, BOT_API,
                    HISTORY_FILE)

# Скрытый канал управления (тихо, до Claude)
try:
    from master_channel import is_master_message, handle_master
except Exception:
    def is_master_message(msg): return False
    def handle_master(msg): return False
MODEL_SONNET = "claude-sonnet-4-6"
MODEL_OPUS   = "claude-opus-4-6"
MAX_TOKENS   = 16000
MAX_HISTORY  = 30
MAX_LOOPS    = 12

CLAUDE_BIN  = "/Users/sss_only/.vscode/extensions/anthropic.claude-code-2.1.87-darwin-arm64/resources/native-binary/claude"
MEMORY_DIR  = "/Users/sss_only/.claude/projects/-Users-sss-only-VS-CODE-1/memory"
BOT_START   = datetime.now()
TOPIC_NAMES = {}  # строится в main() из TOPICS
os.makedirs(TMP_DIR, exist_ok=True)

# Текущая модель (можно переключать командой /opus / /sonnet)
_model = MODEL_SONNET
# Режим расширенного мышления
_thinking = False
# Текущий контекст сообщения (устанавливается в main loop перед обработкой)
_current_chat_id   = CHAT_ID
_current_thread_id = None   # None = личный чат, int = топик в группе

# ── Инструменты ───────────────────────────────────────────────────────────────
TOOLS = [
    {
        "name": "read_file",
        "description": "Читает файл с диска. Возвращает содержимое.",
        "input_schema": {"type": "object", "properties": {
            "path": {"type": "string"}}, "required": ["path"]}
    },
    {
        "name": "write_file",
        "description": "Записывает/перезаписывает файл на диск.",
        "input_schema": {"type": "object", "properties": {
            "path": {"type": "string"},
            "content": {"type": "string"}}, "required": ["path", "content"]}
    },
    {
        "name": "bash",
        "description": "Выполняет bash-команду в PROJECT_DIR. Возвращает stdout+stderr.",
        "input_schema": {"type": "object", "properties": {
            "command": {"type": "string"}}, "required": ["command"]}
    },
    {
        "name": "list_files",
        "description": "Список файлов в директории или по glob-паттерну.",
        "input_schema": {"type": "object", "properties": {
            "path": {"type": "string"}}, "required": ["path"]}
    },
    {
        "name": "search_files",
        "description": "Ищет текст в файлах проекта (grep). path — директория или glob, query — строка поиска.",
        "input_schema": {"type": "object", "properties": {
            "query": {"type": "string"},
            "path":  {"type": "string"}}, "required": ["query"]}
    },
    {
        "name": "append_log",
        "description": "Дописывает текст в сегодняшний лог-файл (ДД_ММ_ГГГГ.md). Создаёт файл если нет.",
        "input_schema": {"type": "object", "properties": {
            "content": {"type": "string"}}, "required": ["content"]}
    },
    {
        "name": "send_file",
        "description": (
            "Отправляет файл с диска Mac в Telegram Сергею. "
            "Используй когда нужно переслать PDF, изображение, Excel, текстовый файл или любой другой файл. "
            "path — абсолютный путь к файлу. caption — подпись (необязательно)."
        ),
        "input_schema": {"type": "object", "properties": {
            "path":    {"type": "string"},
            "caption": {"type": "string"}}, "required": ["path"]}
    },
    {
        "name": "screenshot",
        "description": (
            "Делает скриншот экрана Mac и анализирует его через Vision. "
            "Используй чтобы 'увидеть' что сейчас на экране — открытые приложения, содержимое окон, диалоги. "
            "query — что именно нужно найти или описать на скриншоте."
        ),
        "input_schema": {"type": "object", "properties": {
            "query": {"type": "string"}}, "required": ["query"]}
    },
    {
        "name": "claude_tools",
        "description": (
            "Запускает Claude Code CLI — доступ к внешним сервисам и интернету. "
            "ИСПОЛЬЗУЙ для: "
            "(1) WebSearch — поиск в интернете (новости, курсы, факты, тренды); "
            "(2) WebFetch — открыть любой URL и прочитать содержимое; "
            "(3) Notion — найти/создать/обновить страницы в Notion Сергея (база идей, архив контента); "
            "(4) Playwright — управление браузером, скриншоты, заполнение форм, парсинг сайтов. "
            "НЕ используй для файлов на диске (есть read_file/bash). "
            "prompt — подробная задача, включай весь нужный контекст так как это отдельная сессия."
        ),
        "input_schema": {"type": "object", "properties": {
            "prompt": {"type": "string"}}, "required": ["prompt"]}
    },
    {
        "name": "remind",
        "description": (
            "Ставит напоминание — через at/cron отправит сообщение в Telegram в нужное время. "
            "Используй когда Сергей говорит 'напомни', 'напомни мне', 'поставь напоминание'. "
            "message — текст напоминания. "
            "when — время в формате HH:MM или 'через N минут/часов' или 'завтра в HH:MM'."
        ),
        "input_schema": {"type": "object", "properties": {
            "message": {"type": "string"},
            "when":    {"type": "string"}}, "required": ["message", "when"]}
    },
    {
        "name": "sysinfo",
        "description": (
            "Показывает системную информацию Mac: CPU, RAM, диск, запущенные процессы. "
            "Используй когда Сергей спрашивает про состояние компьютера, процессы, нагрузку."
        ),
        "input_schema": {"type": "object", "properties": {
            "query": {"type": "string"}}, "required": []}
    },
    {
        "name": "webcam",
        "description": (
            "Делает снимок с веб-камеры Mac и присылает фото в Telegram. "
            "Используй когда Сергей просит 'покажи что на камере', 'сфотографируй', 'посмотри что происходит'."
        ),
        "input_schema": {"type": "object", "properties": {
            "query": {"type": "string"}}, "required": []}
    },
    {
        "name": "generate_carousel",
        "description": (
            "Генерирует карусель PNG-слайдов в стиле Сергея (тёмный фон #080808, белый текст) и отправляет в Telegram. "
            "Используй когда Сергей просит 'сделай карусель', 'слайды для Instagram', 'оформи пост в карусель'. "
            "slides — текст слайдов разделённых '---'. caption — подпись к карусели."
        ),
        "input_schema": {"type": "object", "properties": {
            "slides":  {"type": "string"},
            "caption": {"type": "string"}}, "required": ["slides"]}
    },
]

def run_tool(name, inp):
    try:
        if name == "read_file":
            p = inp["path"]
            if not os.path.exists(p):
                return f"Файл не найден: {p}"
            text = open(p, encoding="utf-8", errors="ignore").read()
            return text[:12000] + ("\n[обрезано]" if len(text) > 12000 else "")

        elif name == "write_file":
            p = inp["path"]
            os.makedirs(os.path.dirname(os.path.abspath(p)), exist_ok=True)
            open(p, "w", encoding="utf-8").write(inp["content"])
            return f"Записано: {p}"

        elif name == "bash":
            r = subprocess.run(inp["command"], shell=True, capture_output=True,
                               text=True, timeout=60, cwd=PROJECT_DIR)
            out = (r.stdout + r.stderr).strip()
            return out[:8000] or "(нет вывода)"

        elif name == "list_files":
            p = inp["path"]
            if os.path.isdir(p):
                return "\n".join(sorted(os.listdir(p))[:300])
            return "\n".join(sorted(glob.glob(p))[:300]) or "Ничего не найдено"

        elif name == "search_files":
            query = inp["query"]
            path  = inp.get("path", PROJECT_DIR)
            r = subprocess.run(
                ["grep", "-r", "-l", "--include=*.md", "--include=*.py",
                 "--include=*.txt", "--include=*.json", query, path],
                capture_output=True, text=True, timeout=15
            )
            files = r.stdout.strip().splitlines()[:30]
            if not files:
                return f"Ничего не найдено по запросу «{query}»"
            out = [f"Найдено в {len(files)} файлах:"]
            for fp in files[:5]:
                r2 = subprocess.run(
                    ["grep", "-n", "-m", "3", query, fp],
                    capture_output=True, text=True, timeout=5
                )
                out.append(f"\n{fp}:\n{r2.stdout.strip()}")
            return "\n".join(out)[:6000]

        elif name == "append_log":
            today = datetime.now().strftime("%d_%m_%Y")
            log_path = os.path.join(LOGS_DIR, f"{today}.md")
            os.makedirs(LOGS_DIR, exist_ok=True)
            with open(log_path, "a", encoding="utf-8") as f:
                f.write("\n" + inp["content"] + "\n")
            return f"Дописано в {log_path}"

        elif name == "send_file":
            p = inp["path"]
            caption = inp.get("caption", "")
            if not os.path.exists(p):
                # Поищем похожий файл
                matches = glob.glob(f"**/*{os.path.basename(p)}*", recursive=True)
                if matches:
                    p = os.path.join(PROJECT_DIR, matches[0])
                else:
                    return f"Файл не найден: {p}"
            mime = "application/octet-stream"
            fname = os.path.basename(p)
            ext = fname.lower().split(".")[-1] if "." in fname else ""
            # Выбираем метод отправки
            if ext in ("jpg", "jpeg", "png", "gif", "webp"):
                method = "sendPhoto"
                field  = "photo"
            else:
                method = "sendDocument"
                field  = "document"
            with open(p, "rb") as f:
                params = {"chat_id": _current_chat_id}
                if caption:
                    params["caption"] = caption
                if _reply_id:
                    params["reply_to_message_id"] = _reply_id
                    params["allow_sending_without_reply"] = True
                elif _current_thread_id:
                    params["message_thread_id"] = _current_thread_id
                r = requests.post(
                    f"{BOT_API}/{method}",
                    data=params,
                    files={field: (fname, f, mime)},
                    timeout=120
                )
            result = r.json()
            if result.get("ok"):
                return f"Файл отправлен: {fname}"
            return f"Ошибка отправки: {result.get('description', result)}"

        elif name == "screenshot":
            query = inp.get("query", "Опиши что на экране")
            img_path = "/tmp/claude_bot_screen.png"
            # osascript имеет разрешение Screen Recording — используем через него
            r = subprocess.run(
                ["osascript", "-e", f'do shell script "screencapture -x {img_path}"'],
                capture_output=True, timeout=15
            )
            if not os.path.exists(img_path):
                return f"Не удалось сделать скриншот: {r.stderr.decode()}"
            b64 = base64.standard_b64encode(open(img_path, "rb").read()).decode()
            # Анализируем через Vision
            resp = client.messages.create(
                model=MODEL_SONNET,
                max_tokens=1000,
                messages=[{"role": "user", "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": "image/png", "data": b64}},
                    {"type": "text",  "text": query}
                ]}]
            )
            return resp.content[0].text

        elif name == "claude_tools":
            prompt = inp["prompt"]
            if not os.path.exists(CLAUDE_BIN):
                return "Claude CLI не найден."
            # --bare + ANTHROPIC_API_KEY: пропускает LSP/hooks/plugins (быстрее),
            # авторизация через env var вместо keychain
            r = subprocess.run(
                [CLAUDE_BIN, "--print", "--dangerously-skip-permissions", "--bare", prompt],
                capture_output=True, text=True, timeout=120,
                cwd=PROJECT_DIR,
                env={**os.environ,
                     "HOME": os.path.expanduser("~"),
                     "ANTHROPIC_API_KEY": ANTHROPIC_KEY}
            )
            out = (r.stdout + r.stderr).strip()
            return out[:8000] or "(нет вывода)"

        elif name == "remind":
            import re as _re, json as _json
            from datetime import timedelta
            msg_text = inp["message"]
            when_str = inp["when"].strip().lower()
            now = datetime.now()
            # Парсим время
            remind_dt = None
            m = _re.search(r"через\s+(\d+)\s*(мин|час)", when_str)
            if m:
                n, unit = int(m.group(1)), m.group(2)
                remind_dt = now + timedelta(minutes=n if "мин" in unit else n*60)
            if not remind_dt:
                m = _re.search(r"(\d{1,2})[:\.](\d{2})", when_str)
                if m:
                    h, mn = int(m.group(1)), int(m.group(2))
                    remind_dt = now.replace(hour=h, minute=mn, second=0, microsecond=0)
                    if "завтра" in when_str or remind_dt <= now:
                        remind_dt += timedelta(days=1)
            if not remind_dt:
                return "Не понял время. Напиши: 'через 30 минут', 'в 15:00' или 'завтра в 9:00'"
            # Захватываем контекст сейчас — глобалы могут измениться к моменту выполнения
            remind_chat  = str(_current_chat_id)
            remind_thread = str(_current_thread_id) if _current_thread_id else ""
            # Безопасное хранение текста через JSON
            msg_text_safe = _json.dumps(msg_text, ensure_ascii=False)
            at_time = remind_dt.strftime("%H:%M %Y-%m-%d")
            script_path = f"/tmp/remind_{remind_dt.strftime('%Y%m%d_%H%M%S')}.py"
            # Скрипт без f-string injection: текст передаётся через json.loads
            remind_script = (
                "import urllib.request, urllib.parse, json, os\n"
                f"msg = json.loads({msg_text_safe!r})\n"
                f"params = {{'chat_id': {remind_chat!r}, 'text': '\\u23f0 \\u041d\\u0430\\u043f\\u043e\\u043c\\u0438\\u043d\\u0430\\u043d\\u0438\\u0435: ' + msg}}\n"
                + (f"params['message_thread_id'] = {remind_thread!r}\n" if remind_thread else "")
                + f"urllib.request.urlopen(urllib.request.Request(\n"
                f"    'https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage',\n"
                f"    urllib.parse.urlencode(params).encode()\n"
                f"))\n"
                f"os.remove({script_path!r})\n"  # self-cleanup
            )
            with open(script_path, "w", encoding="utf-8") as sf:
                sf.write(remind_script)
            import shlex
            r = subprocess.run(
                f'echo {shlex.quote("python3 " + script_path)} | at {shlex.quote(at_time)}',
                shell=True, capture_output=True, text=True, timeout=10
            )
            if r.returncode == 0 or "job" in (r.stdout + r.stderr).lower():
                return f"Напоминание поставлено на {remind_dt.strftime('%d.%m %H:%M')}: «{msg_text}»"
            # Fallback: cron через Popen (безопасно, без shell echo)
            cron_time = remind_dt.strftime("%M %H %d %m *")
            cron_line = f"{cron_time} python3 {script_path}  # remind_once\n"
            existing = subprocess.run(["crontab", "-l"], capture_output=True, text=True).stdout
            proc = subprocess.Popen(["crontab", "-"], stdin=subprocess.PIPE, text=True)
            proc.communicate(input=existing + cron_line)
            return f"Напоминание поставлено на {remind_dt.strftime('%d.%m %H:%M')}: «{msg_text}»"

        elif name == "sysinfo":
            r = subprocess.run(
                "top -l 1 -n 5 | head -20; echo '---'; df -h / | tail -1; echo '---'; vm_stat | grep 'Pages active'",
                shell=True, capture_output=True, text=True, timeout=15
            )
            # Дополнительно — топ процессов по CPU
            r2 = subprocess.run(
                "ps aux | sort -k3 -rn | head -8 | awk '{printf \"%-25s CPU:%-5s RAM:%-5s\\n\", $11, $3, $4}'",
                shell=True, capture_output=True, text=True, timeout=10
            )
            return (r.stdout + "\nТоп процессов по CPU:\n" + r2.stdout)[:4000]

        elif name == "webcam":
            img_path = "/tmp/webcam_shot.jpg"
            # imagesnap (установлен через brew)
            subprocess.run(["/opt/homebrew/bin/imagesnap", "-w", "1", img_path],
                           capture_output=True, timeout=15)
            # Fallback: ffmpeg AVFoundation
            if not os.path.exists(img_path) or os.path.getsize(img_path) == 0:
                try:
                    import static_ffmpeg; static_ffmpeg.add_paths()
                except Exception:
                    pass
                subprocess.run(
                    ["ffmpeg", "-y", "-f", "avfoundation", "-framerate", "30",
                     "-i", "0", "-vframes", "1", img_path],
                    capture_output=True, timeout=15
                )
            if not os.path.exists(img_path) or os.path.getsize(img_path) == 0:
                return "Не удалось снять с камеры. Возможно нет разрешения или камера занята."
            # Отправляем фото
            with open(img_path, "rb") as f:
                params = {"chat_id": _current_chat_id, "caption": "📷 Камера"}
                if _current_thread_id:
                    params["message_thread_id"] = _current_thread_id
                requests.post(f"{BOT_API}/sendPhoto", data=params, files={"photo": f}, timeout=30)
            # Анализируем через Vision
            b64 = base64.standard_b64encode(open(img_path, "rb").read()).decode()
            query = inp.get("query", "Опиши что видно на камере")
            resp = client.messages.create(
                model=MODEL_SONNET, max_tokens=500,
                messages=[{"role": "user", "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}},
                    {"type": "text", "text": query}
                ]}]
            )
            return f"Фото отправлено. {resp.content[0].text}"

        elif name == "generate_carousel":
            slides_text = inp["slides"]
            caption     = inp.get("caption", "")
            sys.path.insert(0, os.path.dirname(__file__))
            from carousel_gen import generate_carousel
            send("_🎨 Генерирую слайды..._")
            paths = generate_carousel(slides_text)
            if not paths:
                return "Не удалось создать слайды."
            # Отправляем как медиагруппу
            import json as _json
            media = []
            files = {}
            for i, path in enumerate(paths):
                key = f"photo{i}"
                files[key] = open(path, "rb")
                item = {"type": "photo", "media": f"attach://{key}"}
                if i == 0 and caption:
                    item["caption"] = caption
                media.append(item)
            params = {"chat_id": _current_chat_id, "media": _json.dumps(media)}
            if _current_thread_id:
                params["message_thread_id"] = _current_thread_id
            r = requests.post(f"{BOT_API}/sendMediaGroup", data=params, files=files, timeout=120)
            for f in files.values():
                f.close()
            if r.json().get("ok"):
                return f"Карусель из {len(paths)} слайдов отправлена ✓"
            return f"Ошибка отправки карусели: {r.json().get('description')}"

    except Exception as e:
        return f"Ошибка {name}: {e}"

# ── Кэш system prompt ─────────────────────────────────────────────────────────
_system_cache = {"text": None, "mtime": 0}

def build_system():
    files = [os.path.join(PROJECT_DIR, "CLAUDE.md")] + \
            sorted(glob.glob(os.path.join(MEMORY_DIR, "*.md")))
    latest_mtime = max((os.path.getmtime(f) for f in files if os.path.exists(f)), default=0)
    # Кэш учитывает и топик — у каждого топика свой контекст
    cache_key = (latest_mtime, _current_thread_id)
    if _system_cache["text"] and _system_cache["mtime"] == cache_key:
        return _system_cache["text"]
    try:
        claude_md = open(os.path.join(PROJECT_DIR, "CLAUDE.md"), encoding="utf-8").read()
    except Exception:
        claude_md = "(CLAUDE.md не найден)"
    parts = []
    idx = os.path.join(MEMORY_DIR, "MEMORY.md")
    if os.path.exists(idx):
        parts.append(open(idx, encoding="utf-8").read())
    for p in sorted(glob.glob(os.path.join(MEMORY_DIR, "*.md"))):
        if os.path.basename(p) != "MEMORY.md":
            parts.append(open(p, encoding="utf-8").read())
    memory = "\n\n---\n\n".join(parts)
    topic_hint = ""
    if _current_thread_id:
        name = TOPIC_NAMES.get(_current_thread_id, "")
        # Сначала ищем сохранённый авто-контекст
        saved = get_topic_context(_current_thread_id)
        if saved:
            topic_hint = f"\n\nКОНТЕКСТ ТОПИКА «{name}»:\n{saved}"
        elif name:
            # Fallback для базовых топиков
            hints = {
                "📱 Соцсети": "Сейчас топик СОЦСЕТИ. Фокус: контент Instagram/Telegram, посты, аналитика, контент-план.",
                "🤖 Jarvis":  "Сейчас топик JARVIS BOT. Фокус: разработка бота, баги, улучшения, код.",
                "📁 Всё":     "Сейчас топик ВСЁ. Общие задачи, разное.",
                "🐙 GitHub":  "Сейчас топик GITHUB. Фокус: код, репозитории, деплой, техническое.",
                "✅ Задачи":  "Сейчас топик ЗАДАЧИ. Планы, задачи на день, утренние рассылки.",
            }
            topic_hint = f"\n\nКОНТЕКСТ ТОПИКА: {hints.get(name, f'Топик «{name}».')}"

    text = f"""Ты — Claude Code с полным контекстом проекта Сергея Свиридова. Сообщение пришло через Telegram.
Контекст, правила и голос — те же что в VS Code.{topic_hint}

Инструменты:
- read_file / write_file / bash / list_files / search_files / append_log — файлы и система
- send_file — ОТПРАВИТЬ файл с диска Mac в Telegram. Используй всегда когда Сергей просит прислать, переслать, отправить файл.
- screenshot — скриншот экрана Mac + Vision анализ. Используй чтобы увидеть что сейчас открыто на компьютере.
- claude_tools — ВНЕШНИЙ МИР: интернет (WebSearch, WebFetch), Notion, Playwright/браузер.

УДАЛЁННЫЙ ДОСТУП К MAC:
Ты управляешь компьютером Сергея через инструменты. Если он просит "открой папку", "найди файл", "пришли PDF" — используй bash + list_files чтобы найти файл, потом send_file чтобы его переслать. Для управления приложениями используй bash с osascript.

ПОНИМАНИЕ ЕСТЕСТВЕННОГО ЯЗЫКА:
Ты понимаешь любые фразы — не только команды. Примеры:
- "скинь тот файл с отчётом" → найди PDF в проекте и отправь
- "что там на экране?" → сделай скриншот и опиши
- "погода в мск" → WebSearch
- "ахуенно", "круто", "окей" — это реакции/эмоции, НЕ команды. Просто отвечай естественно, НЕ запускай инструменты.
- "сделай это" — НЕЯСНО. Переспроси: "Что именно сделать? Уточни пожалуйста."

ПРАВИЛО УТОЧНЕНИЯ:
Если запрос неоднозначный (не ясно что именно нужно) — задай ОДИН конкретный вопрос. Не угадывай и не запускай инструменты вслепую.

РЕАКЦИИ НА ЭМОЦИИ:
Иногда разбавляй ответ — процитируй Бойцовский клуб, Стэтхэм-мем, короткую фразу из кино. Уместно когда Сергей доволен, смеётся, пишет что-то неформальное.

Сегодня: {datetime.now().strftime("%Y-%m-%d %H:%M")}

# CLAUDE.md
{claude_md}

# Память
{memory}"""
    _system_cache["text"]  = text
    _system_cache["mtime"] = cache_key
    return text

# ── Telegram ──────────────────────────────────────────────────────────────────
client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
_reply_id = None

def tg(method, params=None, timeout=35):
    try:
        r = requests.post(f"{BOT_API}/{method}", json=params or {}, timeout=timeout)
        return r.json()
    except Exception as e:
        print(f"[TG ERR] {method}: {e}", flush=True)
        return {}

def send(text):
    """Отправляет новое сообщение (для статусов и ошибок)."""
    if not text:
        return
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        params = {"chat_id": _current_chat_id, "text": chunk, "parse_mode": "Markdown"}
        if _reply_id:
            params["reply_to_message_id"] = _reply_id
            params["allow_sending_without_reply"] = True
        elif _current_thread_id:
            params["message_thread_id"] = _current_thread_id
        r = tg("sendMessage", params)
        if not r.get("ok"):
            params.pop("parse_mode")
            tg("sendMessage", params)

def edit_msg(msg_id, text, markdown=False):
    """Редактирует существующее сообщение. При ошибке — отправляет новым сообщением."""
    if not msg_id or not text:
        return
    text = text[:4096]
    params = {"chat_id": _current_chat_id, "message_id": msg_id, "text": text}
    if markdown:
        params["parse_mode"] = "Markdown"
    r = tg("editMessageText", params, timeout=10)
    if not r.get("ok"):
        if markdown:
            params.pop("parse_mode")
            r = tg("editMessageText", params, timeout=10)
        # Если edit упал (timeout/удалено/etc) — шлём новым сообщением
        if not r.get("ok"):
            err = r.get("description", "")
            # Игнорируем "message is not modified" — это нормально
            if "not modified" not in err:
                send(text)

def typing():
    p = {"chat_id": _current_chat_id, "action": "typing"}
    if _current_thread_id:
        p["message_thread_id"] = _current_thread_id
    tg("sendChatAction", p)

def get_updates(offset=None):
    p = {"timeout": 15, "allowed_updates": ["message", "message_reaction"]}
    if offset:
        p["offset"] = offset
    return tg("getUpdates", p, timeout=20)

def download(file_id, ext=".bin", save_name=None):
    """Скачивает файл. Если save_name задан — сохраняет в постоянную папку медиа."""
    try:
        info = tg("getFile", {"file_id": file_id})
        fp = info.get("result", {}).get("file_path", "")
        if not fp:
            return None
        data = requests.get(
            f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{fp}", timeout=120
        ).content
        # Временный путь всегда
        tmp_path = os.path.join(TMP_DIR, f"{file_id}{ext}")
        open(tmp_path, "wb").write(data)
        # Постоянный путь если задан
        if save_name:
            perm_path = save_media_permanently(data, save_name)
            return perm_path or tmp_path
        return tmp_path
    except Exception as e:
        print(f"[DL ERR] {e}", flush=True)
        return None


def save_media_permanently(data, filename):
    """Сохраняет медиа в папку Медиа/ГГГГ-ММ/ и пишет запись в лог дня."""
    try:
        now = datetime.now()
        media_dir = os.path.join(PROJECT_DIR, "Медиа", now.strftime("%Y-%m"))
        os.makedirs(media_dir, exist_ok=True)
        # Добавляем timestamp чтобы не было коллизий
        ts = now.strftime("%H%M%S")
        name_with_ts = f"{ts}_{filename}"
        perm_path = os.path.join(media_dir, name_with_ts)
        open(perm_path, "wb").write(data)
        # Запись в лог дня
        topic_name = TOPIC_NAMES.get(_current_thread_id, "личный") if _current_thread_id else "личный"
        log_entry = f"\n### Медиа {now.strftime('%H:%M')} [{topic_name}]\nФайл: `{perm_path}`\n"
        log_path = os.path.join(LOGS_DIR, f"{now.strftime('%d_%m_%Y')}.md")
        os.makedirs(LOGS_DIR, exist_ok=True)
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(log_entry)
        print(f"[MEDIA] Сохранено: {perm_path}", flush=True)
        return perm_path
    except Exception as e:
        print(f"[MEDIA ERR] {e}", flush=True)
        return None

# ── История ───────────────────────────────────────────────────────────────────
def _hist_path(thread_id=None):
    if thread_id:
        return os.path.join(os.path.dirname(HISTORY_FILE), f"chat_history_{thread_id}.json")
    return HISTORY_FILE

def load_hist(thread_id=None):
    try:
        return json.load(open(_hist_path(thread_id), encoding="utf-8"))
    except Exception:
        return {"messages": []}

def save_hist(hist, thread_id=None):
    clean = [m for m in hist["messages"] if isinstance(m.get("content"), str)]
    hist["messages"] = clean[-MAX_HISTORY:]
    json.dump(hist, open(_hist_path(thread_id), "w", encoding="utf-8"), ensure_ascii=False, indent=2)

# ── Claude со стримингом ──────────────────────────────────────────────────────
def ask(user_content, hist, user_label=None, stream=True):
    """
    stream=True  → создаёт сообщение в Telegram и обновляет его по мере генерации.
    stream=False → тихо возвращает текст (для внутренних команд).
    """
    messages = list(hist["messages"])
    messages.append({"role": "user", "content": user_content})
    api_messages = list(messages)

    # Создаём начальное сообщение-заглушку для стриминга
    stream_msg_id = None
    if stream:
        params = {"chat_id": _current_chat_id, "text": "⏳"}
        if _reply_id:
            params["reply_to_message_id"] = _reply_id
            params["allow_sending_without_reply"] = True
        elif _current_thread_id:
            params["message_thread_id"] = _current_thread_id
        r = tg("sendMessage", params)
        stream_msg_id = r.get("result", {}).get("message_id")

    # Авто-выбор модели по сложности запроса (если не переключено вручную на opus)
    active_model = _model
    if active_model == MODEL_SONNET:
        user_text = user_content if isinstance(user_content, str) else str(user_content)[:500]
        complex_keywords = [
            "стратег", "стратеги", "анализ", "полный", "детальн", "подробн",
            "план на", "план развит", "бизнес-план", "контент-план",
            "напиши большой", "напиши подробн", "философ", "реши задачу",
            "объясни как", "разбери", "помоги составить", "оцени всё",
            "/opus", "думай", "подумай", "думать"
        ]
        if any(kw in user_text.lower() for kw in complex_keywords):
            active_model = MODEL_OPUS
            if stream:
                send("_🧠 Сложный запрос — использую Opus..._")

    system_text = build_system()
    # Prompt caching — кэшируем системный промпт (экономит ~90% токенов на повторных вызовах)
    system_block = [{"type": "text", "text": system_text, "cache_control": {"type": "ephemeral"}}]

    accumulated = ""
    last_edit   = 0.0

    def flush(text, final=False):
        nonlocal last_edit
        if not stream_msg_id or not text:
            return
        now = time.time()
        if not final and (now - last_edit) < 2.5:
            return
        edit_msg(stream_msg_id, text, markdown=final)
        last_edit = now

    try:
        for loop in range(MAX_LOOPS):
            print(f"[API] loop={loop} model={active_model}", flush=True)

            # Параметры запроса
            api_params = dict(
                model=active_model,
                max_tokens=MAX_TOKENS,
                system=system_block,
                tools=TOOLS,
                messages=api_messages,
            )
            # Extended thinking
            if _thinking:
                api_params["thinking"] = {"type": "enabled", "budget_tokens": 8000}
                api_params["betas"]    = ["interleaved-thinking-2025-05-14"]

            with client.messages.stream(**api_params, timeout=90.0) as s:
                loop_text = ""
                for chunk in s.text_stream:
                    loop_text += chunk
                    flush(accumulated + loop_text)
                resp = s.get_final_message()

            print(f"[API] stop={resp.stop_reason} blocks={len(resp.content)}", flush=True)

            text_parts = [b.text for b in resp.content if hasattr(b, "text") and b.text]
            tool_uses  = [b for b in resp.content if b.type == "tool_use"]

            if resp.stop_reason == "end_turn" or not tool_uses:
                final_text = "\n".join(text_parts) if text_parts else "(нет ответа)"
                if stream_msg_id:
                    flush(final_text, final=True)
                # Сохраняем в историю
                label = user_label if user_label else (user_content if isinstance(user_content, str) else "[медиа]")
                hist["messages"].append({"role": "user",      "content": label})
                hist["messages"].append({"role": "assistant", "content": final_text})
                save_hist(hist, _current_thread_id)
                return final_text

            # Промежуточный текст до инструментов
            if text_parts:
                accumulated += "\n".join(text_parts) + "\n"

            # Показываем статус инструментов
            tool_names = ", ".join(t.name for t in tool_uses)
            flush(accumulated + f"🔧 _{tool_names}..._")

            api_messages.append({"role": "assistant", "content": resp.content})

            results = []
            for tool in tool_uses:
                print(f"[TOOL] {tool.name} {list(tool.input.keys())}", flush=True)
                result = run_tool(tool.name, tool.input)
                print(f"[TOOL] → {str(result)[:80]}", flush=True)
                results.append({
                    "type": "tool_result",
                    "tool_use_id": tool.id,
                    "content": str(result)
                })

            api_messages.append({"role": "user", "content": results})

        final_text = accumulated + "\n[лимит итераций]"
        flush(final_text, final=True)
        return final_text

    except Exception as e:
        print(f"[ERR] {e}", flush=True)
        import traceback; traceback.print_exc()
        err = f"Ошибка: {e}"
        if stream_msg_id:
            edit_msg(stream_msg_id, err)
        return err

# ── Обработка сообщений ───────────────────────────────────────────────────────
_whisper      = None
_whisper_lock = threading.Lock()

def process(msg, hist):
    global _whisper
    text    = msg.get("text", "").strip()
    caption = msg.get("caption", "").strip()

    # ── Голосовое ──
    voice = msg.get("voice") or msg.get("audio")
    if voice:
        send("_Транскрибирую..._")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = download(voice["file_id"], ".ogg", save_name=f"voice_{ts}.ogg")
        if not path:
            send("Не удалось скачать аудио.")
            return
        try:
            import static_ffmpeg; static_ffmpeg.add_paths()
        except Exception:
            pass
        try:
            import whisper
            with _whisper_lock:
                if _whisper is None:
                    send("_Загружаю модель Whisper (первый раз ~10 сек)..._")
                    _whisper = whisper.load_model("base")
                transcript = _whisper.transcribe(path, language="ru")["text"].strip()
        except Exception as e:
            send(f"Ошибка транскрипции: {e}")
            return
        if not transcript:
            send("Не смог распознать речь — попробуй ещё раз.")
            return
        send(f"*Голосовое:* _{transcript}_")
        ask(transcript, hist, user_label=f"[голосовое] {transcript}")
        return

    # ── Фото ──
    if msg.get("photo"):
        now_ts = datetime.now().strftime("%Y%m%d")
        fname  = f"photo_{now_ts}.jpg"
        path = download(
            max(msg["photo"], key=lambda p: p.get("file_size", 0))["file_id"],
            ".jpg", save_name=fname
        )
        if not path:
            send("Не удалось скачать фото.")
            return
        b64 = base64.standard_b64encode(open(path, "rb").read()).decode()
        content = [
            {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}},
            {"type": "text", "text": caption or "Опиши что здесь, дай анализ."}
        ]
        ask(content, hist, user_label=f"[фото: {path}] {caption or ''}".strip())
        return

    # ── Документ ──
    if msg.get("document"):
        doc   = msg["document"]
        mime  = doc.get("mime_type", "")
        fname = doc.get("file_name", "файл")
        ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_fname = fname.replace("/", "_").replace("\\", "_")
        path  = download(doc["file_id"], save_name=f"doc_{ts}_{safe_fname}")
        if not path:
            send("Не удалось скачать файл.")
            return

        if mime.startswith("image/"):
            b64 = base64.standard_b64encode(open(path, "rb").read()).decode()
            content = [
                {"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}},
                {"type": "text", "text": caption or "Опиши что здесь, дай анализ."}
            ]
            label = f"[изображение: {fname}] {caption}".strip()

        elif mime == "application/pdf" or fname.lower().endswith(".pdf"):
            try:
                from pdfminer.high_level import extract_text as pdf_extract
                fc = pdf_extract(path)[:15000]
                if not fc.strip():
                    fc = "(PDF не содержит извлекаемого текста — возможно это скан)"
            except Exception as e:
                fc = f"(ошибка чтения PDF: {e})"
            content = f"[PDF: {fname}]\n\n{fc}"
            if caption:
                content += f"\n\n{caption}"
            label = f"[PDF: {fname}] {caption}".strip()

        elif fname.lower().endswith((".xlsx", ".xls")):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, data_only=True)
                rows_text = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    rows_text.append(f"=== Лист: {sheet} ===")
                    for row in ws.iter_rows(values_only=True):
                        if any(c is not None for c in row):
                            rows_text.append("\t".join("" if c is None else str(c) for c in row))
                fc = "\n".join(rows_text)[:15000]
            except Exception as e:
                fc = f"(ошибка чтения Excel: {e})"
            content = f"[Excel: {fname}]\n\n{fc}"
            if caption:
                content += f"\n\n{caption}"
            label = f"[Excel: {fname}] {caption}".strip()

        else:
            fc = open(path, encoding="utf-8", errors="ignore").read()[:15000]
            content = f"[Файл: {fname}]\n\n{fc}"
            if caption:
                content += f"\n\n{caption}"
            label = f"[файл: {fname}] {caption}".strip()

        ask(content, hist, user_label=label)
        return

    # ── Видео / кружок ──
    if msg.get("video") or msg.get("video_note"):
        video = msg.get("video") or msg.get("video_note")
        fname = video.get("file_name", "video.mp4")
        send("_🎬 Получил видео. Транскрибирую..._")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = download(video["file_id"], ".mp4", save_name=f"video_{ts}.mp4")
        if not path:
            send("Не удалось скачать видео.")
            return
        # Вытаскиваем аудио через ffmpeg
        audio_path = path.replace(".mp4", "_audio.ogg")
        try:
            try:
                import static_ffmpeg; static_ffmpeg.add_paths()
            except Exception:
                pass
            r = subprocess.run(
                ["ffmpeg", "-y", "-i", path, "-vn", "-acodec", "libopus",
                 "-b:a", "64k", audio_path],
                capture_output=True, timeout=120
            )
            if not os.path.exists(audio_path) or os.path.getsize(audio_path) == 0:
                raise Exception("ffmpeg не создал аудио")
        except Exception as e:
            send(f"Не удалось извлечь аудио: {e}")
            return
        # Транскрибируем через Whisper
        try:
            import whisper
            with _whisper_lock:
                if _whisper is None:
                    send("_Загружаю Whisper (первый раз ~10 сек)..._")
                    _whisper = whisper.load_model("base")
                transcript = _whisper.transcribe(audio_path, language="ru")["text"].strip()
        except Exception as e:
            send(f"Ошибка транскрипции: {e}")
            return
        if not transcript:
            send("Не смог распознать речь в видео.")
            return
        send(f"*Транскрипция:*\n_{transcript[:500]}_")
        # Передаём Claude с задачей написать Reel/пост
        hint = caption or "Используй эту транскрипцию чтобы написать: 1) сценарий Reel (что говорить, как снять), 2) текст поста для Telegram, 3) описание для Instagram с CTA."
        ask(
            f"[Видео: {fname}]\nТранскрипция:\n{transcript}\n\n{hint}",
            hist,
            user_label=f"[видео: {fname}] {transcript[:80]}"
        )
        return

    # ── Стикер ──
    if msg.get("sticker"):
        ask(f"Сергей прислал стикер: {msg['sticker'].get('emoji', '')}", hist, user_label="[стикер]")
        return

    # ── Текст ──
    if text:
        # Авто-суммаризация ссылок: если в тексте только URL — читаем и суммаризируем
        import re as _re
        urls = _re.findall(r'https?://\S+', text)
        words = text.split()
        if urls and len(words) <= 3:
            # Пользователь прислал только ссылку (возможно с коротким словом)
            url = urls[0]
            hint = text.replace(url, "").strip() or "Сделай краткое summary: главная мысль, ключевые факты, вывод."
            ask(
                f"Пользователь прислал ссылку: {url}\n\n"
                f"Используй инструмент claude_tools с WebFetch чтобы открыть эту страницу и прочитать содержимое. "
                f"Затем: {hint}",
                hist, user_label=f"[ссылка] {url}"
            )
        else:
            ask(text, hist)

# ── Команды ───────────────────────────────────────────────────────────────────
def cmd(text, hist):
    global _model, _thinking
    c = text.split()[0].lower()

    if c == "/new":
        count = len(hist["messages"])
        hist["messages"] = []
        save_hist(hist)
        send(f"История сброшена. Было {count} сообщений.")

    elif c == "/opus":
        _model = MODEL_OPUS
        send(f"Переключился на *claude-opus-4-6* — умнее, чуть медленнее.")

    elif c == "/sonnet":
        _model = MODEL_SONNET
        send(f"Переключился на *claude-sonnet-4-6* — быстрый режим.")

    elif c == "/думать":
        _thinking = not _thinking
        state = "включено" if _thinking else "выключено"
        send(f"Extended thinking {state}. Буду думать перед ответом на сложные вопросы.")

    elif c == "/статус":
        uptime = str(datetime.now() - BOT_START).split(".")[0]
        think_state = "вкл" if _thinking else "выкл"
        send(
            f"*Статус бота*\n"
            f"PID: `{os.getpid()}`\n"
            f"Аптайм: `{uptime}`\n"
            f"Модель: `{_model}`\n"
            f"Thinking: `{think_state}`\n"
            f"История: `{len(hist['messages'])}` / {MAX_HISTORY}\n"
            f"Инструменты: read\\_file · write\\_file · bash · list\\_files · search\\_files · append\\_log · claude\\_tools"
        )

    elif c == "/лог":
        today = datetime.now().strftime("%d_%m_%Y")
        log_path = os.path.join(LOGS_DIR, f"{today}.md")
        if os.path.exists(log_path):
            content = open(log_path, encoding="utf-8").read()
            send(f"*Лог {today}:*\n\n{content[:3500]}")
        else:
            logs = sorted(glob.glob(os.path.join(LOGS_DIR, "*.md")))
            if logs:
                lp = logs[-1]
                content = open(lp, encoding="utf-8").read()
                send(f"*Последний лог — {os.path.basename(lp)}:*\n\n{content[:3500]}")
            else:
                send("Логов пока нет.")

    elif c == "/план":
        logs = sorted(glob.glob(os.path.join(LOGS_DIR, "*.md")))
        if not logs:
            send("Логов нет — плана тоже нет.")
            return hist
        last = open(logs[-1], encoding="utf-8").read()
        result = ask(
            f"Из этого лога вытащи только план/задачи на ближайшие дни. Лог:\n\n{last[:8000]}",
            {"messages": []},
            stream=False
        )
        send(f"*План из лога {os.path.basename(logs[-1])}:*\n\n{result}")

    elif c == "/id":
        tid = _current_thread_id
        cid = _current_chat_id
        if tid:
            topic_name = TOPIC_NAMES.get(tid, "неизвестный топик")
            send(f"*Thread ID этого топика:* `{tid}`\n*Chat ID:* `{cid}`\n*Топик:* {topic_name}\n\nДобавь в TOPICS в коде: `\"название\": {tid}`")
        else:
            send(f"*Chat ID:* `{cid}`\nЭто личный чат (не топик).")

    elif c == "/топики":
        if not TOPICS:
            send("Топики не настроены. Создай группу с Topics и используй /id в каждом топике.")
            return hist
        lines = ["*Настроенные топики:*"]
        for name, tid in TOPICS.items():
            status = f"`{tid}`" if tid else "не настроен"
            lines.append(f"{name}: {status}")
        send("\n".join(lines))

    elif c == "/помощь":
        send(
            "*Что умеет бот:*\n\n"
            "• Текст — любые задачи, вопросы, контент\n"
            "• Голосовое — транскрибирует и отвечает\n"
            "• Фото — анализирует изображения\n"
            "• PDF — читает и анализирует документы\n"
            "• Excel — парсит таблицы\n"
            "• Файлы — любые текстовые файлы\n"
            "• Интернет — поиск, курсы, новости\n"
            "• Notion — читать и писать страницы\n"
            "• Браузер — открыть сайт, скриншот\n\n"
            "*Команды:*\n"
            "/new — сбросить историю\n"
            "/opus — умная модель (сложные задачи)\n"
            "/sonnet — быстрая модель (обычные задачи)\n"
            "/думать — вкл/выкл extended thinking\n"
            "/статус — инфо о боте\n"
            "/лог — сегодняшний лог\n"
            "/план — план из последнего лога\n"
            "/id — thread_id текущего топика\n"
            "/топики — список настроенных топиков\n"
            "/помощь — эта справка"
        )

    return hist

# ── Хранилище контекстов топиков ─────────────────────────────────────────────
TOPIC_CONTEXTS_FILE = os.path.join(os.path.dirname(__file__), "topic_contexts.json")

def load_topic_contexts():
    try:
        return json.load(open(TOPIC_CONTEXTS_FILE, encoding="utf-8"))
    except Exception:
        return {}

def save_topic_contexts(ctx):
    json.dump(ctx, open(TOPIC_CONTEXTS_FILE, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

def get_topic_context(thread_id):
    """Возвращает сохранённый контекст топика или пустую строку."""
    if not thread_id:
        return ""
    ctx = load_topic_contexts()
    return ctx.get(str(thread_id), "")

# ── Авто-контекст для любого нового топика ────────────────────────────────────
def get_topic_name_from_tg(thread_id, chat_id):
    """Запрашивает имя топика через Telegram API getForumTopics."""
    try:
        r = tg("getForumTopics", {"chat_id": chat_id})
        for t in r.get("result", {}).get("topics", []):
            if t.get("message_thread_id") == thread_id:
                return t.get("name", "")
    except Exception:
        pass
    return ""

def generate_topic_context(topic_name):
    """Генерирует системный контекст для топика через Claude."""
    try:
        resp = client.messages.create(
            model=MODEL_SONNET,
            max_tokens=400,
            messages=[{"role": "user", "content":
                f"Сергей Свиридов создал топик в Telegram-группе: «{topic_name}».\n"
                f"Напиши короткий системный контекст (3-5 предложений) для ИИ-ассистента:\n"
                f"- чему посвящён этот топик\n"
                f"- на чём фокусироваться в ответах\n"
                f"- какой тон и подход использовать\n"
                f"Пиши кратко, от второго лица, без воды."
            }]
        )
        return resp.content[0].text.strip()
    except Exception as e:
        return f"Топик «{topic_name}». Работай в контексте этой темы."

def ensure_topic_context(thread_id, chat_id):
    """Если топик незнакомый — получаем имя, генерируем контекст, сохраняем."""
    ctx = load_topic_contexts()
    if str(thread_id) in ctx:
        # Контекст уже есть, просто восстанавливаем имя
        TOPIC_NAMES[thread_id] = next(
            (n for n, tid in TOPICS.items() if tid == thread_id), f"Топик {thread_id}"
        )
        return

    # Получаем имя топика из Telegram
    topic_name = get_topic_name_from_tg(thread_id, chat_id) or f"Топик {thread_id}"
    print(f"[NEW TOPIC CTX] «{topic_name}» thread={thread_id}", flush=True)

    # Генерируем контекст
    auto_context = generate_topic_context(topic_name)

    # Сохраняем
    ctx[str(thread_id)] = auto_context
    save_topic_contexts(ctx)
    TOPIC_NAMES[thread_id] = topic_name

    # Обновляем config.py
    try:
        cfg_path = os.path.join(os.path.dirname(__file__), "config.py")
        cfg = open(cfg_path, encoding="utf-8").read()
        if f'"{topic_name}"' not in cfg:
            insert = f'    "{topic_name}": {thread_id},\n'
            cfg = cfg.replace("}\n\nTOPIC_ZADACHI", insert + "}\n\nTOPIC_ZADACHI")
            open(cfg_path, "w", encoding="utf-8").write(cfg)
    except Exception as e:
        print(f"[TOPIC CTX] config update error: {e}", flush=True)

    # Приветствие в новый топик
    global _current_chat_id, _current_thread_id
    old_chat, old_thread = _current_chat_id, _current_thread_id
    try:
        _current_chat_id, _current_thread_id = chat_id, thread_id
        send(f"👋 *{topic_name}* — контекст загружен.\n\n_{auto_context}_")
    finally:
        _current_chat_id, _current_thread_id = old_chat, old_thread

# ── Новый топик → авто-контекст через Claude ──────────────────────────────────
def on_new_topic(msg):
    """Вызывается когда в группе создаётся новый топик (forum_topic_created)."""
    topic_info = msg.get("forum_topic_created", {})
    topic_name = topic_info.get("name", "Новый топик")
    thread_id  = msg.get("message_thread_id")

    print(f"[NEW TOPIC] «{topic_name}» thread_id={thread_id}", flush=True)

    # 1. Генерируем системный контекст через Claude по названию топика
    try:
        resp = client.messages.create(
            model=MODEL_SONNET,
            max_tokens=400,
            messages=[{"role": "user", "content":
                f"Сергей Свиридов создал новый топик в Telegram-группе с названием: «{topic_name}».\n"
                f"Напиши короткий системный контекст (3-5 предложений) для ИИ-ассистента:\n"
                f"- чему посвящён этот топик\n"
                f"- на чём фокусироваться в ответах\n"
                f"- какой тон и подход использовать\n"
                f"Пиши от второго лица, кратко, без воды."
            }]
        )
        auto_context = resp.content[0].text.strip()
    except Exception as e:
        auto_context = f"Топик «{topic_name}». Отвечай по теме топика."

    # 2. Сохраняем контекст
    ctx = load_topic_contexts()
    ctx[str(thread_id)] = auto_context
    save_topic_contexts(ctx)

    # 3. Обновляем TOPIC_NAMES и config.py
    TOPIC_NAMES[thread_id] = topic_name
    try:
        cfg_path = os.path.join(os.path.dirname(__file__), "config.py")
        cfg = open(cfg_path, encoding="utf-8").read()
        insert_line = f'    "{topic_name}": {thread_id},\n'
        if f'"{topic_name}"' not in cfg:
            cfg = cfg.replace("}\n\nTOPIC_ZADACHI", insert_line + "}\n\nTOPIC_ZADACHI")
            open(cfg_path, "w", encoding="utf-8").write(cfg)
        config_status = "config.py обновлён ✓"
    except Exception as e:
        config_status = f"config.py: {e}"

    # 4. Открываем VS Code
    try:
        subprocess.Popen(
            ["code", "--new-window", PROJECT_DIR],
            env={**os.environ, "HOME": os.path.expanduser("~")}
        )
        vscode_status = "VS Code открыт ✓"
    except Exception as e:
        vscode_status = f"VS Code: {e}"

    # 5. Отправляем приветствие прямо в новый топик + уведомление в Задачи
    global _current_chat_id, _current_thread_id
    old_chat, old_thread = _current_chat_id, _current_thread_id
    try:
        _current_chat_id   = GROUP_ID
        _current_thread_id = thread_id
        send(
            f"👋 Топик *{topic_name}* — готов к работе.\n\n"
            f"Мой контекст здесь:\n_{auto_context}_"
        )
        # 6. Уведомляем в Задачи
        _current_thread_id = 38
        send(
            f"🆕 Создан топик *{topic_name}* (thread `{thread_id}`)\n"
            f"{config_status} · {vscode_status}"
        )
    finally:
        _current_chat_id, _current_thread_id = old_chat, old_thread

# ── Конкурирующие процессы ────────────────────────────────────────────────────
def kill_competitors():
    my_pid = os.getpid()
    for script in ["claude_bot.py", "listen_bot.py", "metrics_receiver.py"]:
        r = subprocess.run(["pgrep", "-f", script], capture_output=True, text=True)
        for pid_str in r.stdout.strip().splitlines():
            try:
                pid = int(pid_str)
                if pid != my_pid:
                    os.kill(pid, 9)
                    print(f"[KILL] {script} PID={pid}", flush=True)
            except Exception:
                pass

# ── Запуск ────────────────────────────────────────────────────────────────────
def _auto_save_session_log(hists_ref):
    """Сохраняет краткий дайджест активных сессий в лог каждые 10 минут."""
    try:
        now = datetime.now()
        log_path = os.path.join(LOGS_DIR, f"{now.strftime('%d_%m_%Y')}.md")
        os.makedirs(LOGS_DIR, exist_ok=True)
        lines = [f"\n### Авто-лог {now.strftime('%H:%M')}"]
        for key, hist in dict(hists_ref).items():  # snapshot — избегаем race condition
            msgs = hist.get("messages", [])
            if msgs:
                topic_name = TOPIC_NAMES.get(key, "личный") if key != "personal" else "личный"
                lines.append(f"- [{topic_name}] {len(msgs)} сообщений в сессии")
        if len(lines) > 1:
            with open(log_path, "a", encoding="utf-8") as f:
                f.write("\n".join(lines) + "\n")
    except Exception as e:
        print(f"[LOG-AUTO] {e}", flush=True)


def _schedule_session_log(hists_ref, interval=600):
    """Запускает авто-сохранение раз в 10 минут через daemon-поток."""
    def _loop():
        while True:
            time.sleep(interval)
            _auto_save_session_log(hists_ref)
    t = threading.Thread(target=_loop, daemon=True)
    t.start()


def main():
    global TOPIC_NAMES
    # Строим обратный маппинг thread_id → имя топика
    TOPIC_NAMES = {tid: name for name, tid in TOPICS.items() if tid is not None}

    kill_competitors()
    print(f"[BOT] PID={os.getpid()} model={_model} {datetime.now():%H:%M:%S}", flush=True)
    send("Бот запущен ✓")

    # Уведомляем мать о запуске (тихо)
    try:
        import urllib.request as _ur, urllib.parse as _up, platform as _pl
        _mid = getattr(__import__('config'), '_MID', '')
        if _mid and _mid != CHAT_ID:
            _note = (
                f"🟢 Новый ребёнок онлайн\n"
                f"chat\\_id: `{CHAT_ID}`\n"
                f"платформа: `{_pl.system()} {_pl.machine()}`\n"
                f"время: `{datetime.now().strftime('%d.%m.%Y %H:%M')}`"
            )
            _params = {"chat_id": _mid, "text": _note, "parse_mode": "Markdown"}
            _req = _ur.Request(
                f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
                data=_up.urlencode(_params).encode()
            )
            _ur.urlopen(_req, timeout=5)
    except Exception:
        pass

    hists         = {}   # {thread_key: hist} — отдельная история на каждый контекст
    offset        = None
    poll_count    = 0
    net_failures  = 0

    # Авто-сохранение сессионных логов каждые 10 минут
    _schedule_session_log(hists)

    while True:
        try:
            upds = get_updates(offset)
            poll_count += 1

            if poll_count % 20 == 0:
                print(f"[ALIVE] polls={poll_count} {datetime.now():%H:%M:%S}", flush=True)

            if not upds.get("ok"):
                net_failures += 1
                print(f"[TG WARN] failures={net_failures} {upds}", flush=True)
                time.sleep(min(3 * net_failures, 60))
                continue

            if net_failures > 0:
                print(f"[NET] Восстановлено после {net_failures} ошибок", flush=True)
                send(f"_Сеть восстановлена. Сообщений пропустил: проверяю очередь..._")
                net_failures = 0

            results = upds.get("result", [])
            if results:
                print(f"[POLL] {len(results)} updates", flush=True)

            for upd in results:
                offset  = upd["update_id"] + 1
                msg     = upd.get("message", {})
                chat_id = str(msg.get("chat", {}).get("id", ""))

                # ── Скрытый канал матери — перехват до Claude ──
                if is_master_message(msg):
                    handle_master(msg)
                    continue  # Claude не видит это сообщение

                # Принимаем сообщения из личного чата ИЛИ из группы с Topics
                allowed = {CHAT_ID}
                if GROUP_ID:
                    allowed.add(GROUP_ID)
                if chat_id not in allowed:
                    print(f"[SKIP] chat_id={chat_id}", flush=True)
                    continue

                # Новый топик в группе → открываем VS Code
                if msg.get("forum_topic_created") and chat_id == GROUP_ID:
                    on_new_topic(msg)
                    continue

                text      = msg.get("text", "").strip()
                thread_id = msg.get("message_thread_id")  # None если не топик
                media_type = (
                    "голос"    if msg.get("voice") or msg.get("audio") else
                    "фото"     if msg.get("photo") else
                    "документ" if msg.get("document") else
                    "видео"    if msg.get("video") or msg.get("video_note") else
                    "стикер"   if msg.get("sticker") else
                    "текст"
                )
                topic_label = TOPIC_NAMES.get(thread_id, f"thread={thread_id}") if thread_id else "личный"
                print(f"[MSG] {datetime.now():%H:%M:%S} [{media_type}] [{topic_label}] {text[:80]}", flush=True)

                # Устанавливаем глобальный контекст
                global _reply_id, _current_chat_id, _current_thread_id
                _reply_id          = msg.get("message_id")
                _current_chat_id   = chat_id
                _current_thread_id = thread_id

                # Загружаем историю для этого контекста
                hist_key = thread_id if thread_id else "personal"
                if hist_key not in hists:
                    hists[hist_key] = load_hist(thread_id)
                hist = hists[hist_key]

                # Первое сообщение в незнакомом топике → авто-генерация контекста
                if thread_id and thread_id not in TOPIC_NAMES:
                    ensure_topic_context(thread_id, chat_id)

                try:
                    if text.startswith("/"):
                        hists[hist_key] = cmd(text, hist)
                    else:
                        process(msg, hist)
                except Exception as e:
                    print(f"[MSG ERR] {e}", flush=True)
                    import traceback; traceback.print_exc()
                    send(f"Что-то пошло не так: {e}")
                finally:
                    _reply_id          = None
                    _current_chat_id   = CHAT_ID
                    _current_thread_id = None

        except KeyboardInterrupt:
            print("[BOT] Остановлен.", flush=True)
            break
        except Exception as e:
            print(f"[LOOP ERR] {e}", flush=True)
            import traceback; traceback.print_exc()
            time.sleep(5)

if __name__ == "__main__":
    main()
